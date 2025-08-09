import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_PATH = "/workspace/I_Bond_Tracker.xlsx"


def set_header(ws, row, headers):
    bold = Font(bold=True)
    fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
    thin = Side(style="thin", color="FFCCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
        cell.fill = fill
        cell.border = border


def autosize(ws):
    for column_cells in ws.columns:
        length = 0
        col = column_cells[0].column_letter
        for cell in column_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            if len(val) > length:
                length = len(val)
        ws.column_dimensions[col].width = min(max(length + 2, 10), 60)


def build_instructions(ws):
    ws.title = "Instructions"
    lines = [
        "I Bond Inventory & Monthly Value Tracker",
        "",
        "How to use:",
        "1) Fill the 'Rates' sheet with May/Nov announcement dates and the published annual VARIABLE rate for each period.",
        "   - Example Effective Dates: 2024-05-01, 2024-11-01, etc.",
        "   - Enter Variable rate as a decimal (e.g., 0.0490 for 4.90%).",
        "2) Enter each bond in the 'Inventory' sheet:",
        "   - BondID: your identifier",
        "   - IssueDate: the bond's issue date (any day in the issue month is ok)",
        "   - PurchaseAmount: the amount you paid (face value)",
        "   - FixedRate: the bond's fixed rate at issue (decimal)",
        "   - MonthsToProject: how many months of schedule to build in 'BondSchedule'",
        "3) To view a monthly schedule for any bond:",
        "   - Go to 'BondSchedule', choose a Bond ID from the dropdown in B1.",
        "   - Columns A–E will spill a monthly schedule with the applicable rate and value.",
        "",
        "Notes:",
        "- This workbook uses modern Excel dynamic array functions (LET, SEQUENCE, SCAN, LAMBDA).",
        "  Use Excel 365 or Excel 2021+ for best results.",
        "- 'CurrentValue' in Inventory estimates accrued value to today using the same model.",
        "- Estimated redemption value (with early redemption penalty) is not included,",
        "  but can be derived similarly by removing roughly the last 3 months of interest",
        "  if held less than 5 years.",
        "- Composite rate per period is computed as: fixed + variable + fixed*variable (all annual).",
        "  Monthly growth within a 6-month period uses the 6th-root of (1 + composite/2).",
    ]
    for r, text in enumerate(lines, start=1):
        ws.cell(row=r, column=1, value=text)
    ws.column_dimensions['A'].width = 110


def build_rates(ws):
    ws.title = "Rates"
    headers = ["EffDate", "VariableRate"]
    set_header(ws, 1, headers)

    # Example placeholders (commented values): leave blank for user input
    example_rows = [
        (datetime.date(2024, 5, 1), 0.0),
        (datetime.date(2024, 11, 1), 0.0),
        (datetime.date(2025, 5, 1), 0.0),
    ]
    for idx, (eff, var) in enumerate(example_rows, start=2):
        ws.cell(row=idx, column=1, value=eff)
        ws.cell(row=idx, column=2, value=var)
        ws.cell(row=idx, column=1).number_format = numbers.FORMAT_DATE_YYYYMMDD2
        ws.cell(row=idx, column=2).number_format = "0.00%"

    # Add a table to Rates
    last_row = max(100, len(example_rows) + 20)
    tab = Table(displayName="tblRates", ref=f"A1:B{last_row}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    autosize(ws)


def build_inventory(ws):
    ws.title = "Inventory"
    headers = [
        "BondID", "Owner", "IssueDate", "PurchaseAmount", "FixedRate", "MonthsToProject", 
        "MonthsHeld", "CurrentValue"
    ]
    set_header(ws, 1, headers)

    # Example row
    example_row = ["EX-0001", "Sample Owner", datetime.date(2024, 1, 15), 1000.00, 0.009, 120]
    for col, value in enumerate(example_row, start=1):
        cell = ws.cell(row=2, column=col, value=value)
        if col == 3:  # IssueDate
            cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
        if col == 4:  # PurchaseAmount
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        if col == 5:  # FixedRate
            cell.number_format = "0.00%"

    # Apply formats to columns
    for r in range(2, 102):
        ws.cell(row=r, column=3).number_format = numbers.FORMAT_DATE_YYYYMMDD2  # IssueDate
        ws.cell(row=r, column=4).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # PurchaseAmount
        ws.cell(row=r, column=5).number_format = "0.00%"  # FixedRate

    # Add computed columns formulas for first 100 rows
    for r in range(2, 102):
        # MonthsHeld
        ws.cell(row=r, column=7, value=f"=IFERROR(DATEDIF(C{r}, TODAY(), \"m\"), \"\")")

        # CurrentValue using dynamic arrays and LET
        # This formula uses:
        # - Rates!A:A as EffDate, Rates!B:B as VariableRate
        # - Composite = fixed + var + fixed*var
        # - Monthly model: value = principal * product(1 + composite/2 for each full 6-mo period)
        #                  * (1 + current composite/2)^(months_into_period/6)
        formula = (
            f"=IF(OR(C{r}=\"\",D{r}=\"\",E{r}=\"\"),\"\","
            f"LET(issue,C{r}, principal,D{r}, fixed,E{r}, months, IFERROR(DATEDIF(issue, TODAY(), \"m\"),0),"
            f" startIdx, MATCH(issue, Rates!A:A, 1), fullPeriods, INT(months/6), monthsInPeriod, MOD(months,6),"
            f" varVec, IF(fullPeriods=0, \"\", INDEX(Rates!B:B, SEQUENCE(fullPeriods,1, startIdx, 1))),"
            f" fullProd, IF(fullPeriods=0, 1, EXP(SUM(LN(1 + ((fixed + varVec + fixed*varVec)/2))))),"
            f" lastVar, INDEX(Rates!B:B, startIdx + fullPeriods),"
            f" partial, POWER(1 + ((fixed + lastVar + fixed*lastVar)/2), monthsInPeriod/6),"
            f" IFERROR(ROUND(principal * fullProd * partial, 2), \"\")))"
        )
        ws.cell(row=r, column=8, value=formula)
        ws.cell(row=r, column=8).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Add table for the Inventory input area (first 100 rows)
    tab = Table(displayName="tblBonds", ref="A1:H101")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    autosize(ws)


def build_bond_schedule(ws):
    ws.title = "BondSchedule"
    ws["A1"] = "Select Bond ID:"

    # Data validation dropdown for BondID
    dv = DataValidation(type="list", formula1="=Inventory!$A$2:$A$101", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(ws["B1"])  # Selected BondID

    # Helper lookups
    ws["A3"] = "IssueDate"
    ws["B3"] = "=IFERROR(XLOOKUP($B$1, Inventory!$A$2:$A$101, Inventory!$C$2:$C$101), \"\")"
    ws["A4"] = "PurchaseAmount"
    ws["B4"] = "=IFERROR(XLOOKUP($B$1, Inventory!$A$2:$A$101, Inventory!$D$2:$D$101), \"\")"
    ws["A5"] = "FixedRate"
    ws["B5"] = "=IFERROR(XLOOKUP($B$1, Inventory!$A$2:$A$101, Inventory!$E$2:$E$101), \"\")"
    ws["A6"] = "MonthsToProject"
    ws["B6"] = "=IFERROR(XLOOKUP($B$1, Inventory!$A$2:$A$101, Inventory!$F$2:$F$101), 0)"

    # Formats
    ws["B3"].number_format = numbers.FORMAT_DATE_YYYYMMDD2
    ws["B4"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws["B5"].number_format = "0.00%"

    # Table headers for the spilled schedule
    headers = ["Month", "MonthIndex", "VariableRate", "CompositeRate", "Value"]
    set_header(ws, 8, headers)

    # Spilled sequences and calculations using dynamic arrays (Excel 365)
    # Month (dates)
    ws["A9"] = (
        "=IF($B$3=\"\" , \"\", EDATE($B$3, SEQUENCE($B$6+1, 1, 0, 1)))"
    )
    ws["A9"].number_format = numbers.FORMAT_DATE_YYYYMMDD2

    # MonthIndex
    ws["B9"] = (
        "=IF($B$3=\"\" , \"\", SEQUENCE($B$6+1, 1, 0, 1))"
    )

    # VariableRate (array depending on MonthIndex)
    ws["C9"] = (
        "=IF($B$3=\"\", \"\", LET(m, B9#, issue, $B$3, startIdx, MATCH(issue, Rates!A:A, 1), "
        "periodIdx, INT(m/6), annIdx, startIdx + periodIdx, INDEX(Rates!B:B, annIdx)))"
    )
    ws["C9"].number_format = "0.00%"

    # CompositeRate array
    ws["D9"] = (
        "=IF($B$3=\"\", \"\", LET(m, B9#, fixed, $B$5, var, C9#, fixed + var + fixed*var))"
    )
    ws["D9"].number_format = "0.00%"

    # Value array via cumulative products per period using SCAN
    ws["E9"] = (
        "=IF($B$3=\"\", \"\", "
        "LET(m, B9#, principal, $B$4, fixed, $B$5, maxM, $B$6, issue, $B$3, "
        "startIdx, MATCH(issue, Rates!A:A, 1), numPeriods, INT(maxM/6)+1, "
        "periodIdxVec, SEQUENCE(numPeriods,1,0,1), annIdxVec, startIdx + periodIdxVec, "
        "varVec, INDEX(Rates!B:B, annIdxVec), halfFacVec, 1 + (fixed + varVec + fixed*varVec)/2, "
        "cumProdVec, SCAN(1, halfFacVec, LAMBDA(a,x,a*x)), "
        "k, INT(m/6), r, MOD(m,6), "
        "fullProd, INDEX(cumProdVec, k+1), partial, POWER(INDEX(halfFacVec, k+1), r/6), "
        "ROUND(principal * fullProd * partial, 2)))"
    )
    ws["E9"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    autosize(ws)


def main():
    wb = Workbook()

    # Build sheets
    build_instructions(wb.active)
    build_rates(wb.create_sheet())
    build_inventory(wb.create_sheet())
    build_bond_schedule(wb.create_sheet())

    # Set a pleasant tab order
    wb._sheets = [
        wb["Instructions"],
        wb["Inventory"],
        wb["Rates"],
        wb["BondSchedule"],
    ]

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    main()